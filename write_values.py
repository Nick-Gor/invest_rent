from openpyxl import load_workbook
import os
from flask import Flask, render_template, request
from markupsafe import escape
from DBcm import UseDatabase
import json

# wb = load_workbook('./invest_rent/Data/meters_Dudina.xlsx')
wb = load_workbook('./Data/meters_Dudina.xlsx')
# sheet = wb['electricity']

def drop_none(cell_:str):
    if cell_ == "None":
        return ""
    return cell_

def parce_excel_val(dict_, sheet):
    meters_dic = {}
    for i in range (dict_['start_l'], dict_['stop_l']):
        value_ = {}
        place_ = []
        value_['enviroment'] = dict_['env_']
        value_['date'] = sheet.cell(row=2, column=dict_['r_date']).value
        value_['month'] = sheet.cell(row=2, column=dict_['r_month']).value
        n = int(sheet.cell(row=i, column=dict_['r_date']).value) 
        p = int(sheet.cell(row=i, column=(dict_['r_date']-2)).value)
        k = int(sheet.cell(row=i, column=dict_['r_coef']).value)
        value_['debet'] =(n-p)*k
        value_['value'] = int(sheet.cell(row=i, column=dict_['r_date']).value)
        meters_dic[str(sheet.cell(row=i, column=dict_['r_meter']).value)] = value_
    return meters_dic

el_meters = {'start_l':3, 'stop_l':129, 'r_date':26, 'r_month':25,
             'r_meter':6, 'r_coef':7, 'env_':'electricity'}
water_meters = {'start_l':3, 'stop_l':51, 'r_date':27, 'r_month':26,
                'r_meter':5, 'r_coef':6, 'env_':'water'}

data = parce_excel_val(el_meters, wb['electricity'])
data.update(parce_excel_val(water_meters, wb['water']))

# f = open("./database/insert_valuesdud.txt", "w", encoding='utf-8')
# for k, v in data.items():
#     f.write(f"('{k}'")
#     for ln in v.values():
#         f.write(f", '{ln}'")
#     f.write(f"),\n")
# f.close()


# i = 1
# for key, value in data.items():
#     print(f"{i}. {key}: {value}")
#     i +=1


dbconfig = {'host': '127.0.0.1',
            'user': 'engineer.sever', 'password': '1234567',
            'database': 'investrent'}

def write_db (key_, dict_) -> None:
    with UseDatabase(dbconfig) as cursor:
        _SQL = """insert into values_( number, enviroment, month, date, debet, value) 
        values (%s, %s, %s, %s, %s, %s)"""
        cursor.execute(_SQL, (key_, dict_['enviroment'], dict_['month'], dict_['date'], dict_['debet'], dict_['value']))


for key_, value in data.items():
    write_db(key_, value)

# with open('values.json', 'w', encoding='utf-8') as json_file:
#     json.dump(data, json_file, ensure_ascii=False, indent=4)
