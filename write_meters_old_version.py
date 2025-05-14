from openpyxl import load_workbook
import os
from flask import Flask, render_template, request
from markupsafe import escape
from DBcm import UseDatabase
import json

# wb = load_workbook('./invest_rent/Data/meters_Dudina.xlsx')
wb = load_workbook('./Data/meters_Dudina.xlsx')
# sheet = wb['electricity']

def drop_none(cell_:str):
    if cell_ == "None":
        return ""
    return cell_

def parce_excel(dict_, sheet):
    meters_dic = {}
    for i in range (dict_['start_l'], dict_['stop_l']):
        value_ = {}
        place_ = []
        for r in range (dict_['start_r'], dict_['stop_r']):
            place_.append(drop_none(str(sheet.cell(row=i, column=r).value)))
        value_['place'] = " ".join(place_)
        value_['coefficient'] = str(sheet.cell(row=i, column=dict_['r_coef']).value)
        value_['tenant'] = str(sheet.cell(row=i, column=dict_['r_org']).value)
        value_['enviroment'] = dict_['env_']
        meters_dic[str(sheet.cell(row=i, column=dict_['r_meter']).value)] = value_
    return meters_dic

el_meters = {'start_l':3, 'stop_l':130, 'start_r':2,'stop_r':5, 'r_org':5,
             'r_meter':6, 'r_coef':7, 'env_':'electricity'}
water_meters = {'start_l':3, 'stop_l':51, 'start_r':2,'stop_r':4, 'r_org':4,
                'r_meter':5,'r_coef':6, 'env_':'water'}

data = parce_excel(el_meters, wb['electricity'])
data.update(parce_excel(water_meters, wb['water']))

# i = 1
# for key, value in data.items():
#     print(f"{i}. {key}: {value}")
#     i +=1
# dbconfig = {'host': '127.0.0.1',
#             'user': 'engineer.sever', 'password': '1234567',
#             'database': 'investrent'}

# def write_db (i, key_, dict_) -> None:
#     with UseDatabase(dbconfig) as cursor:
#         _SQL = """insert into metersdud(num, number, coefficient, place, tenant, enviroment) 
#         values (%s, %s, %s, %s, %s, %s)"""
#         cursor.execute(_SQL, (i, key_, dict_['coefficient'], dict_['place'], dict_['tenant'], dict_['enviroment']))

# i = 1
# for key_, value in data.items():
#     write_db(i, key_, value)
#     i += 1

with open('meters.json', 'w', encoding='utf-8') as json_file:
    json.dump(data, json_file, ensure_ascii=False, indent=4)